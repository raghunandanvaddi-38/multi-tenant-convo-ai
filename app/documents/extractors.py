"""
Text extractors — one function per supported format.

Return plain text. Best-effort; empty string on failure (caller marks the doc
failed via the exception message).
"""

from __future__ import annotations

import csv
import io
import logging

import PyPDF2
import docx


log = logging.getLogger("documents.extractors")


def extract_txt(data: bytes) -> str:
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_markdown(data: bytes) -> str:
    return extract_txt(data)


def extract_pdf(data: bytes) -> str:
    reader = PyPDF2.PdfReader(io.BytesIO(data))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception as e:
            raise ValueError(f"encrypted PDF ({e})")
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def extract_docx(data: bytes) -> str:
    doc = docx.Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def extract_csv(data: bytes) -> str:
    text = extract_txt(data)
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        lines.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(lines)


def extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("XLSX support requires 'openpyxl'. Install with: pip install openpyxl") from e
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))
    return "\n".join(parts)


def extract_html(data: bytes) -> str:
    text = extract_txt(data)
    # Strip tags without adding beautifulsoup as a hard dep
    import re
    text = re.sub(r"<script.*?</script>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<style.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


EXTRACTORS = {
    ".txt": extract_txt,
    ".md": extract_markdown,
    ".markdown": extract_markdown,
    ".pdf": extract_pdf,
    ".docx": extract_docx,
    ".csv": extract_csv,
    ".xlsx": extract_xlsx,
    ".html": extract_html,
    ".htm": extract_html,
}


def extract(filename: str, data: bytes) -> str:
    import os
    ext = os.path.splitext(filename.lower())[1]
    fn = EXTRACTORS.get(ext)
    if fn is None:
        raise ValueError(f"Unsupported file type: {ext or '(no extension)'}")
    return fn(data)


def supported_extensions() -> list[str]:
    return sorted(EXTRACTORS.keys())
